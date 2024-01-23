import os
import openpyxl as pyxl
import datetime import date

class tracker: 

    def initialize(tracker_loc)
    if not os.path.exists(tracker_loc):
        wb= pyxl.Workbook()
        ws=wb.active
        ws1= wb.create_sheet('Data_Tracking',0)
        header=['Date updated', 'feature class', 'Number of new features', 'Source', 'Updated by']
        header_style=pyxl.styles.Font(size=12, bold=True)
        ws1=wb['Data_Tracking']
        ws1.append(header)
        for cell in ws1["1:1"]:
            cell.font=header_style
        wb.save(tracker_loc)
    else:
        wb= pyxl.load_workbook(tracker_loc)
        ws1=wb['Data_Tracking']

    def append_tracker(fc_name,feats,method,username)
    xlsx_append=[date.today, fc_name, feats, method, username]
            print(xlsx_append)
            ws1.append(xlsx_append)
    